from fastapi import APIRouter, HTTPException, BackgroundTasks, Request
from typing import Dict, List, Optional
import logging
import uuid
import json
from datetime import datetime
import pandas as pd
import io
from fastapi.responses import StreamingResponse

from models_turnos import (
    Configuracion, Ciclo, Posicion, Operador, AsignacionDiaria,
    SolicitudOptimizacionTurnos, ResultadoOptimizacionTurnos,
    EstadisticasOptimizacion, TipoTurno, EstadoOperador
)
from optimizer_turnos import OptimizadorTurnos
from optimizer_heuristica import optimizar_con_heuristica
import os

logger = logging.getLogger(__name__)

# Router para endpoints de turnos
router = APIRouter(prefix="/turnos", tags=["turnos"])

# Almacenamiento en memoria para resultados de optimización
resultados_optimizacion: Dict[str, ResultadoOptimizacionTurnos] = {}
tareas_en_progreso: Dict[str, str] = {}

@router.post("/debug/", response_model=dict)
async def debug_datos(request: Request):
    """
    Endpoint temporal para debug - recibe datos sin validación
    """
    try:
        datos = await request.json()
        print("=== DATOS RAW RECIBIDOS ===")
        print(json.dumps(datos, indent=2, ensure_ascii=False))
        print("=== FIN DATOS RAW ===")
        return {"mensaje": "Datos recibidos correctamente", "datos": datos}
    except Exception as e:
        print(f"Error al procesar datos: {e}")
        return {"error": str(e)}

@router.post("/optimizar/", response_model=Dict[str, str])
async def optimizar_turnos(
    solicitud: SolicitudOptimizacionTurnos,
    background_tasks: BackgroundTasks
):
    """
    Inicia la optimización de turnos de forma asíncrona
    """
    try:
        # Logging detallado para debug
        print("=== DATOS RECIBIDOS EN /turnos/optimizar/ ===")
        print(f"Tipo de solicitud: {type(solicitud)}")
        print(f"Configuración: {solicitud.configuracion}")
        print(f"Ciclos: {solicitud.ciclos}")
        print(f"Posiciones: {solicitud.posiciones}")
        print(f"Operadores: {solicitud.operadores}")
        print("=== FIN DATOS RECIBIDOS ===")
        
        # Generar ID único para la tarea
        tarea_id = str(uuid.uuid4())
        
        # Marcar tarea como en progreso
        tareas_en_progreso[tarea_id] = "iniciando"
        
        # Ejecutar optimización en background
        background_tasks.add_task(ejecutar_optimizacion, tarea_id, solicitud)
        
        logger.info(f"Optimización iniciada con ID: {tarea_id}")
        
        return {
            "tarea_id": tarea_id,
            "estado": "iniciado",
            "mensaje": "Optimización iniciada. Use el endpoint /estado/{tarea_id} para verificar el progreso."
        }
        
    except Exception as e:
        logger.error(f"Error al iniciar optimización: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al iniciar optimización: {str(e)}")

@router.get("/estado/{tarea_id}")
async def obtener_estado_optimizacion(tarea_id: str):
    """
    Obtiene el estado de una optimización en progreso
    """
    try:
        # Verificar si la tarea está en progreso
        if tarea_id in tareas_en_progreso:
            return {
                "tarea_id": tarea_id,
                "estado": tareas_en_progreso[tarea_id],
                "completado": False,
                "mensaje": f"Optimización en progreso: {tareas_en_progreso[tarea_id]}"
            }
        
        # Verificar si la tarea está completada
        if tarea_id in resultados_optimizacion:
            resultado = resultados_optimizacion[tarea_id]
            return {
                "tarea_id": tarea_id,
                "estado": "completado",
                "completado": True,
                "factible": resultado.factible,
                "mensaje": resultado.mensaje,
                "estadisticas": resultado.estadisticas,
                "errores": resultado.errores,
                "advertencias": resultado.advertencias
            }
        
        # Tarea no encontrada
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener estado: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener estado: {str(e)}")

@router.get("/resultado/{tarea_id}")
async def obtener_resultado_optimizacion(tarea_id: str):
    """
    Obtiene el resultado completo de una optimización
    """
    try:
        if tarea_id not in resultados_optimizacion:
            raise HTTPException(status_code=404, detail="Resultado no encontrado")
        
        resultado = resultados_optimizacion[tarea_id]
        
        if not resultado.factible:
            raise HTTPException(status_code=400, detail="La optimización no encontró una solución factible")
        
        return resultado
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener resultado: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener resultado: {str(e)}")

@router.get("/descargar/{tarea_id}")
async def descargar_cronograma_excel(tarea_id: str):
    """
    Descarga el cronograma optimizado como archivo Excel con 4 pestañas
    """
    try:
        if tarea_id not in resultados_optimizacion:
            raise HTTPException(status_code=404, detail="Resultado no encontrado")
        
        resultado = resultados_optimizacion[tarea_id]
        
        if not resultado.factible:
            raise HTTPException(status_code=400, detail="No hay cronograma disponible para descargar")
        
        # Verificar si existe el archivo Excel generado
        if not hasattr(resultado, 'archivo_excel') or not resultado.archivo_excel:
            raise HTTPException(status_code=404, detail="Archivo Excel no generado")
        
        archivo_excel = resultado.archivo_excel
        
        if not os.path.exists(archivo_excel):
            raise HTTPException(status_code=404, detail="Archivo Excel no encontrado en el servidor")
        
        # Leer archivo y enviarlo
        with open(archivo_excel, 'rb') as f:
            contenido = f.read()
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cronograma_turnos_{timestamp}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(contenido),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al descargar cronograma: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al descargar cronograma: {str(e)}")

@router.delete("/resultado/{tarea_id}")
async def eliminar_resultado(tarea_id: str):
    """
    Elimina un resultado de optimización de la memoria
    """
    try:
        eliminado = False
        
        if tarea_id in resultados_optimizacion:
            del resultados_optimizacion[tarea_id]
            eliminado = True
        
        if tarea_id in tareas_en_progreso:
            del tareas_en_progreso[tarea_id]
            eliminado = True
        
        if not eliminado:
            raise HTTPException(status_code=404, detail="Resultado no encontrado")
        
        return {"mensaje": "Resultado eliminado exitosamente"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar resultado: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al eliminar resultado: {str(e)}")

@router.get("/resultados/")
async def listar_resultados():
    """
    Lista todos los resultados disponibles
    """
    try:
        resultados = []
        
        # Agregar tareas en progreso
        for tarea_id, estado in tareas_en_progreso.items():
            resultados.append({
                "tarea_id": tarea_id,
                "estado": estado,
                "completado": False,
                "factible": None
            })
        
        # Agregar resultados completados
        for tarea_id, resultado in resultados_optimizacion.items():
            resultados.append({
                "tarea_id": tarea_id,
                "estado": "completado",
                "completado": True,
                "factible": resultado.factible,
                "mensaje": resultado.mensaje
            })
        
        return {"resultados": resultados}
        
    except Exception as e:
        logger.error(f"Error al listar resultados: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al listar resultados: {str(e)}")

async def ejecutar_optimizacion(tarea_id: str, solicitud: SolicitudOptimizacionTurnos):
    """
    Ejecuta la optimización en background usando el optimizador heurístico
    """
    try:
        logger.info(f"Iniciando optimización heurística para tarea {tarea_id}")
        
        # Actualizar estado
        tareas_en_progreso[tarea_id] = "preparando_datos"
        
        # Actualizar estado
        tareas_en_progreso[tarea_id] = "optimizando"
        
        # Ejecutar optimización HEURÍSTICA
        cronograma, metricas = optimizar_con_heuristica(solicitud)
        
        # Preparar resultado en formato esperado
        resultado_opt = {
            "cronograma": cronograma,
            "estadisticas": metricas,
            "observaciones": []
        }
        
        # Actualizar estado
        tareas_en_progreso[tarea_id] = "generando_excel"
        
        # Generar Excel con las 4 pestañas
        archivo_excel = await _generar_excel_completo(resultado_opt, tarea_id)
        
        # Convertir resultado a formato compatible con ResultadoOptimizacionTurnos
        # (por ahora simplificado, luego se puede mejorar)
        resultado_api = ResultadoOptimizacionTurnos(
            cronograma=[],  # Vacío por ahora, el cronograma está en el Excel
            estadisticas=resultado_opt.get("estadisticas", {}),
            advertencias=resultado_opt.get("observaciones", []),
            errores=[],
            factible=True,
            mensaje=f"Optimización completada. Excel generado: {archivo_excel}"
        )
        
        # Agregar ruta del Excel al resultado
        resultado_api.archivo_excel = archivo_excel
        
        # Guardar resultado
        resultados_optimizacion[tarea_id] = resultado_api
        
        # Remover de tareas en progreso
        if tarea_id in tareas_en_progreso:
            del tareas_en_progreso[tarea_id]
        
        logger.info(f"Optimización completada para tarea {tarea_id}")
        
    except Exception as e:
        logger.error(f"Error en optimización {tarea_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Guardar resultado con error
        resultado_error = ResultadoOptimizacionTurnos(
            cronograma=[],
            estadisticas={},
            advertencias=[],
            errores=[str(e)],
            factible=False,
            mensaje=f"Error en optimización: {str(e)}"
        )
        
        resultados_optimizacion[tarea_id] = resultado_error
        
        # Remover de tareas en progreso
        if tarea_id in tareas_en_progreso:
            del tareas_en_progreso[tarea_id]


def _aplicar_formato_excel(archivo: str, operadores_map: dict):
    """Aplica formato avanzado a las pestañas del Excel"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(archivo)
    
    # ===== PESTAÑA CRONOGRAMA =====
    if "Cronograma" in wb.sheetnames:
        ws_crono = wb["Cronograma"]
        ws_crono.freeze_panes = "A2"
        ws_crono.auto_filter.ref = ws_crono.dimensions
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_crono[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info("✓ Formato aplicado a pestaña 'Cronograma'")
    
    # ===== PESTAÑA CALENDARIO =====
    if "Calendario" in wb.sheetnames:
        ws_cal = wb["Calendario"]
        ws_cal.insert_rows(1)
        
        color_td = PatternFill(start_color="5D8F3E", end_color="5D8F3E", fill_type="solid")
        color_tn = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
        color_de = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        color_vc = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        dias_por_mes = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        
        ws_cal['A1'] = ""
        ws_cal['B1'] = ""
        ws_cal['A2'] = "Operador"
        ws_cal['B2'] = "Posición"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell_ref in ['A2', 'B2']:
            ws_cal[cell_ref].fill = header_fill
            ws_cal[cell_ref].font = header_font
            ws_cal[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        
        col_idx = 3
        for mes_num, (mes_nombre, dias_mes) in enumerate(zip(meses_nombres, dias_por_mes), start=1):
            start_col = col_idx
            end_col = col_idx + dias_mes - 1
            
            ws_cal.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws_cal.cell(row=1, column=start_col)
            cell.value = mes_nombre
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for dia in range(1, dias_mes + 1):
                cell = ws_cal.cell(row=2, column=col_idx)
                cell.value = dia
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1
            
            # Las columnas separadoras ya están en el DataFrame (SEP_XX)
            # Solo avanzar el índice sin agregar nada
            if mes_num < 12:
                col_idx += 1
        
        for row_idx, row in enumerate(ws_cal.iter_rows(min_row=3), start=3):
            for cell in row[2:]:
                if cell.value == "TD":
                    cell.fill = color_td
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "TN":
                    cell.fill = color_tn
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "DE":
                    cell.fill = color_de
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "VC":
                    cell.fill = color_vc
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_cal.freeze_panes = "C3"
        ws_cal.column_dimensions['A'].width = 30
        ws_cal.column_dimensions['B'].width = 12
        
        # Aplicar ancho a columnas de días y separadores
        col_idx = 3
        for mes_num in range(1, 13):
            # Columnas de días del mes
            for _ in range(dias_por_mes[mes_num - 1]):
                col_letter = get_column_letter(col_idx)
                ws_cal.column_dimensions[col_letter].width = 3
                col_idx += 1
            
            # Columna separadora (más ancha para crear espacio visual)
            if mes_num < 12:
                col_letter = get_column_letter(col_idx)
                ws_cal.column_dimensions[col_letter].width = 2
                col_idx += 1
        
        ws_cal.row_dimensions[1].height = 20
        ws_cal.row_dimensions[2].height = 18
        
        logger.info("✓ Formato aplicado a pestaña 'Calendario'")
    
    # ===== PESTAÑA PROGRAMACIÓN =====
    if "Programacion" in wb.sheetnames:
        ws_prog = wb["Programacion"]
        
        colores_operadores = {
            1: PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid"),
            2: PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
            3: PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid"),
            4: PatternFill(start_color="6FA8DC", end_color="6FA8DC", fill_type="solid"),
            5: PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid"),
            6: PatternFill(start_color="E06666", end_color="E06666", fill_type="solid"),
            0: PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        }
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for row_idx, row in enumerate(ws_prog.iter_rows(min_row=1), start=1):
            if row[0].value == "Posición":
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                for cell in row[3:]:
                    if isinstance(cell.value, int):
                        cell.fill = colores_operadores.get(cell.value, colores_operadores[0])
                        cell.font = Font(bold=True, color="000000", size=9)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.value == "":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(0, 3):
                    cell = row[col]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_prog.freeze_panes = "D1"
        ws_prog.column_dimensions['A'].width = 15
        ws_prog.column_dimensions['B'].width = 12
        ws_prog.column_dimensions['C'].width = 8
        
        for col_idx in range(4, ws_prog.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_prog.column_dimensions[col_letter].width = 3
        
        logger.info("✓ Formato aplicado a pestaña 'Programacion'")
    
    # ===== PESTAÑA LEYENDA =====
    if "Leyenda" in wb.sheetnames:
        ws_ley = wb["Leyenda"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_ley[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        colores_operadores = {
            1: PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid"),
            2: PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
            3: PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid"),
            4: PatternFill(start_color="6FA8DC", end_color="6FA8DC", fill_type="solid"),
            5: PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid"),
            6: PatternFill(start_color="E06666", end_color="E06666", fill_type="solid"),
            0: PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        }
        
        for row_idx, row in enumerate(ws_ley.iter_rows(min_row=2), start=2):
            numero_cell = ws_ley.cell(row=row_idx, column=1)
            if isinstance(numero_cell.value, int):
                numero_cell.fill = colores_operadores.get(numero_cell.value, colores_operadores[0])
                numero_cell.font = Font(bold=True, color="000000", size=11)
                numero_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            nombre_cell = ws_ley.cell(row=row_idx, column=2)
            nombre_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        ws_ley.column_dimensions['A'].width = 10
        ws_ley.column_dimensions['B'].width = 35
        
        logger.info("✓ Formato aplicado a pestaña 'Leyenda'")
    
    wb.save(archivo)
    logger.info(f"✓ Formato guardado en: {archivo}")


async def _generar_excel_completo(resultado_opt, tarea_id: str) -> str:
    """
    Genera el archivo Excel con las 4 pestañas usando el cronograma optimizado
    """
    try:
        import pandas as pd
        from datetime import date
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        logger.info("\n" + "🔵" * 30)
        logger.info("GENERANDO EXCEL COMPLETO")
        logger.info(f"Tarea ID: {tarea_id}")
        
        # Convertir cronograma a DataFrame
        df = pd.DataFrame(resultado_opt["cronograma"])
        logger.info(f"Total registros en cronograma: {len(df)}")
        logger.info(f"Columnas del cronograma: {df.columns.tolist()}")
        
        # Mostrar primeros 3 registros como muestra
        logger.info("\n📋 MUESTRA DE CRONOGRAMA (primeros 3 registros):")
        for idx, row in df.head(3).iterrows():
            logger.info(f"  {idx+1}. Fecha={row.get('Fecha')}, Nombre={row.get('Nombre')}, "
                       f"Posición={row.get('Posicion')}, Posición_Inicial={row.get('Posicion_Inicial')}, "
                       f"Estado={row.get('Estado')}, Estado2={row.get('Estado2')}")
        
        # Mapeo de operadores a números para la vista "Programación"
        operadores_unicos = df["Nombre"].unique()
        operadores_map = {nombre: idx + 1 for idx, nombre in enumerate(operadores_unicos) if nombre != "op_vacaciones"}
        
        # ===== GENERAR VISTA CALENDARIO =====
        logger.info("=" * 60)
        logger.info("INICIANDO GENERACIÓN DE VISTA CALENDARIO")
        logger.info(f"Total operadores únicos: {len(operadores_unicos)}")
        logger.info(f"Operadores: {operadores_unicos}")
        
        # 🔧 FIX: Crear columna datetime para ordenamiento cronológico correcto
        df['Fecha_dt'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y')
        
        # Agrupar operadores por posición para mejor organización visual
        operadores_por_posicion = {}
        for nombre in operadores_unicos:
            df_op = df[df["Nombre"] == nombre].sort_values("Fecha_dt")
            if df_op.empty:
                continue
            posicion_inicial = df_op.iloc[0]["Posicion_Inicial"] if "Posicion_Inicial" in df_op.columns else df_op.iloc[0]["Posicion"]
            if posicion_inicial not in operadores_por_posicion:
                operadores_por_posicion[posicion_inicial] = []
            operadores_por_posicion[posicion_inicial].append(nombre)
        
        logger.info(f"Operadores agrupados por posición: {operadores_por_posicion}")
        
        calendario_data = []
        posiciones_ordenadas = sorted([p for p in operadores_por_posicion.keys() if p != "op_vacaciones"]) + (["op_vacaciones"] if "op_vacaciones" in operadores_por_posicion else [])
        
        for idx_pos, posicion in enumerate(posiciones_ordenadas):
            # Agregar separador visual entre grupos (excepto antes del primero)
            if idx_pos > 0:
                separador = {"Operador": "", "Posición": ""}
                calendario_data.append(separador)
                logger.info(f"\n━━━ SEPARADOR VISUAL ━━━")
            
            logger.info(f"\n┌─ GRUPO: {posicion} ─┐")
            
            for nombre in operadores_por_posicion[posicion]:
                # ORDENAR POR Fecha_dt (datetime) NO por Fecha (string alfabético)
                df_op = df[df["Nombre"] == nombre].sort_values("Fecha_dt")
                logger.info(f"\n--- Procesando operador: {nombre} ---")
                logger.info(f"  Registros encontrados: {len(df_op)}")
                
                if df_op.empty:
                    logger.warning(f"  ⚠️ Sin registros para {nombre}, saltando...")
                    continue
                
                # Obtener posición inicial
                posicion_inicial = df_op.iloc[0]["Posicion_Inicial"] if "Posicion_Inicial" in df_op.columns else df_op.iloc[0]["Posicion"]
                logger.info(f"  Posición inicial detectada: {posicion_inicial}")
                
                fila = {"Operador": nombre, "Posición": posicion_inicial}
                
                # Contar estados para debug
                estados_count = df_op["Estado2"].value_counts().to_dict()
                logger.info(f"  Estados encontrados: {estados_count}")
                
                # Log para debug (solo primeros 5 registros)
                primeros_5 = df_op.head(5)
                logger.info(f"  Primeros 5 registros de {nombre}:")
                for idx, row in primeros_5.iterrows():
                    logger.info(f"    idx={idx}, Fecha={row['Fecha']}, Estado2={row['Estado2']}")
                
                # Agregar datos con separadores entre meses
                mes_anterior = None
                for _, row in df_op.iterrows():
                    fecha_str = row["Fecha"]
                    fecha_obj = pd.to_datetime(fecha_str, format="%d/%m/%Y")
                    
                    # Si cambiamos de mes y no es el primer mes, agregar columna separadora
                    if mes_anterior is not None and fecha_obj.month != mes_anterior:
                        col_separador = f"SEP_{mes_anterior:02d}"
                        fila[col_separador] = ""
                    
                    col_nombre = f"{fecha_obj.month:02d}_{fecha_obj.day:02d}"
                    fila[col_nombre] = row["Estado2"]
                    mes_anterior = fecha_obj.month
                
                calendario_data.append(fila)
                logger.info(f"  ✓ Operador {nombre} agregado al calendario ({len(fila)-2} días)")
        
        df_calendario = pd.DataFrame(calendario_data)
        logger.info(f"\n✅ Calendario generado: {len(df_calendario)} operadores")
        logger.info("=" * 60)
        
        # ===== GENERAR VISTA PROGRAMACIÓN =====
        logger.info("\n" + "=" * 60)
        logger.info("INICIANDO GENERACIÓN DE VISTA PROGRAMACIÓN")
        
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        dias_por_mes = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        
        # Obtener posiciones únicas desde la columna Posición (donde realmente trabajaron)
        # Filtrar solo posiciones válidas (no descanso, vacaciones, etc.)
        posiciones_todas = df["Posicion"].unique()
        logger.info(f"Posiciones encontradas en cronograma: {posiciones_todas}")
        
        posiciones_lista = [p for p in posiciones_todas if p not in ["descanso", "vacaciones", "op_vacaciones"]]
        logger.info(f"Posiciones válidas para programación: {posiciones_lista}")
        
        # Análisis de datos por posición
        for pos in posiciones_lista:
            registros_pos = df[df["Posicion"] == pos]
            logger.info(f"\n--- Posición: {pos} ---")
            logger.info(f"  Total registros: {len(registros_pos)}")
            estados = registros_pos["Estado"].value_counts().to_dict()
            logger.info(f"  Estados: {estados}")
            operadores_en_pos = registros_pos["Nombre"].unique()
            logger.info(f"  Operadores que trabajaron aquí: {operadores_en_pos}")
        
        programacion_data = []
        for idx_pos, posicion in enumerate(posiciones_lista):
            logger.info(f"\n🔄 Generando datos para posición: {posicion}")
            if idx_pos > 0:
                # Fila vacía
                fila_vacia = {"Posición": "", "Mes": "", "Turno": ""}
                for dia in range(1, 32):
                    fila_vacia[str(dia)] = ""
                programacion_data.append(fila_vacia)
                
                # Header repetido
                header_row = {"Posición": "Posición", "Mes": "Mes", "Turno": "Turno"}
                for dia in range(1, 32):
                    header_row[str(dia)] = dia
                programacion_data.append(header_row)
            
            # Contador para debug (solo loguear primer mes de cada posición)
            log_detail = True
            
            for mes_idx, (mes_nombre, dias_mes) in enumerate(zip(meses_nombres, dias_por_mes), start=1):
                for turno in ["TD", "TN"]:
                    fila = {"Posición": posicion, "Mes": mes_nombre, "Turno": turno}
                    
                    if log_detail and mes_idx == 1:  # Solo loguear enero
                        logger.info(f"  📅 {mes_nombre} - {turno}:")
                    
                    encontrados = 0
                    vacios = 0
                    
                    for dia in range(1, 32):
                        if dia <= dias_mes:
                            fecha_obj = date(2025, mes_idx, dia)
                            fecha_str = fecha_obj.strftime("%d/%m/%Y")
                            estado_buscar = "t.dia" if turno == "TD" else "t.noche"
                            
                            # Buscar operador que trabajó en esta posición en este día/turno
                            # Solo usar la columna Posición (donde realmente trabajó)
                            registro = df[
                                (df["Posicion"] == posicion) &
                                (df["Fecha"] == fecha_str) &
                                (df["Estado"] == estado_buscar)
                            ]
                            
                            if not registro.empty:
                                nombre_op = registro.iloc[0]["Nombre"]
                                numero_op = operadores_map.get(nombre_op, "")
                                encontrados += 1
                                
                                if log_detail and mes_idx == 1 and dia <= 5:  # Primeros 5 días
                                    logger.info(f"    Día {dia}: Op #{numero_op} ({nombre_op})")
                            else:
                                numero_op = ""
                                vacios += 1
                                
                                if log_detail and mes_idx == 1 and dia <= 5:
                                    logger.info(f"    Día {dia}: VACÍO (no se encontró registro)")
                            
                            fila[str(dia)] = numero_op  # ✅ SIN "D" - solo el número
                        else:
                            fila[str(dia)] = ""  # ✅ SIN "D"
                    
                    if log_detail and mes_idx == 1:
                        logger.info(f"    → Resumen {turno}: {encontrados} asignados, {vacios} vacíos")
                    
                    programacion_data.append(fila)
                
                if log_detail and mes_idx == 1:
                    log_detail = False  # Solo loguear el primer mes
        
        df_programacion = pd.DataFrame(programacion_data)
        logger.info(f"\n✅ Programación generada: {len(df_programacion)} filas")
        logger.info("=" * 60)
        
        # ===== GENERAR LEYENDA =====
        leyenda_data = []
        for nombre, numero in sorted(operadores_map.items(), key=lambda x: x[1]):
            leyenda_data.append({"Número": numero, "Operador": nombre})
        df_leyenda = pd.DataFrame(leyenda_data)
        
        # ===== EXPORTAR A EXCEL =====
        archivo = f"cronograma_{tarea_id}.xlsx"
        
        # Remover columna "Posicion_Inicial" del cronograma (solo para uso interno)
        df_export = df.copy()
        if "Posicion_Inicial" in df_export.columns:
            df_export = df_export.drop(columns=["Posicion_Inicial"])
        
        # ===== LOGS DE VERIFICACIÓN =====
        logger.info("\n" + "🔍" * 30)
        logger.info("VERIFICACIÓN DE DATOS ANTES DE EXPORTAR")
        
        # Verificar datos de AGUIRRE en Cronograma
        aguirre_cronograma = df_export[df_export["Nombre"].str.contains("AGUIRRE", na=False)].head(10)
        logger.info(f"\n📊 AGUIRRE en Cronograma (primeros 10 días):")
        for _, row in aguirre_cronograma.iterrows():
            logger.info(f"  {row['Fecha']}: Posición={row.get('Posicion', 'N/A')}, Estado2={row.get('Estado2', 'N/A')}")
        
        # Verificar datos de AGUIRRE en Calendario
        aguirre_calendario = df_calendario[df_calendario["Operador"].str.contains("AGUIRRE", na=False)]
        if not aguirre_calendario.empty:
            logger.info(f"\n📅 AGUIRRE en Calendario:")
            logger.info(f"  Posición: {aguirre_calendario.iloc[0].get('Posicion', 'N/A')}")
            # Mostrar primeros 10 días
            cols = [c for c in aguirre_calendario.columns if c not in ['Operador', 'Posicion']][:10]
            valores = [aguirre_calendario.iloc[0].get(c, 'N/A') for c in cols]
            for col, val in zip(cols, valores):
                logger.info(f"  {col}: {val}")
        
        logger.info("🔍" * 30 + "\n")
        
        with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name="Cronograma", index=False)
            df_calendario.to_excel(writer, sheet_name="Calendario", index=False)
            df_programacion.to_excel(writer, sheet_name="Programacion", index=False)
            df_leyenda.to_excel(writer, sheet_name="Leyenda", index=False)
        
        # ===== APLICAR FORMATO AVANZADO =====
        _aplicar_formato_excel(archivo, operadores_map)
        
        logger.info(f"Excel generado: {archivo}")
        return archivo
        
    except Exception as e:
        logger.error(f"Error al generar Excel: {str(e)}")
        raise

# Endpoints de configuración y datos maestros

@router.post("/configuracion/validar")
async def validar_configuracion(configuracion: Configuracion):
    """
    Valida una configuración antes de la optimización
    """
    try:
        errores = []
        advertencias = []
        
        # Validaciones básicas
        if configuracion.limite_horas_anuales <= 0:
            errores.append("El límite de horas anuales debe ser mayor a 0")
        
        if configuracion.limite_porcentaje_dia < 0 or configuracion.limite_porcentaje_dia > 100:
            errores.append("El límite de porcentaje día debe estar entre 0 y 100")
        
        if configuracion.limite_porcentaje_noche < 0 or configuracion.limite_porcentaje_noche > 100:
            errores.append("El límite de porcentaje noche debe estar entre 0 y 100")
        
        if configuracion.mes_inicio_analisis < 1 or configuracion.mes_inicio_analisis > 12:
            errores.append("El mes de inicio debe estar entre 1 y 12")
        
        # Advertencias
        if configuracion.limite_horas_anuales > 2500:
            advertencias.append("El límite de horas anuales es muy alto (>2500)")
        
        return {
            "valido": len(errores) == 0,
            "errores": errores,
            "advertencias": advertencias
        }
        
    except Exception as e:
        logger.error(f"Error al validar configuración: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al validar configuración: {str(e)}")

@router.post("/datos/validar")
async def validar_datos_entrada(solicitud: SolicitudOptimizacionTurnos):
    """
    Valida los datos de entrada antes de la optimización
    """
    try:
        errores = []
        advertencias = []
        
        # Validar operadores
        if not solicitud.operadores:
            errores.append("Debe proporcionar al menos un operador")
        
        # Validar posiciones
        if not solicitud.posiciones:
            errores.append("Debe proporcionar al menos una posición")
        
        # Validar ciclos
        if not solicitud.ciclos:
            errores.append("Debe proporcionar al menos un ciclo")
        
        # Validar compatibilidad operadores-posiciones
        tipos_posicion_operadores = set(op.id_tipo_posicion for op in solicitud.operadores)
        tipos_posicion_posiciones = set(pos.tipo_posicion for pos in solicitud.posiciones)
        
        if not tipos_posicion_operadores.intersection(tipos_posicion_posiciones):
            errores.append("No hay operadores compatibles con las posiciones definidas")
        
        # Verificar cobertura mínima
        for tipo_pos in tipos_posicion_posiciones:
            operadores_tipo = [op for op in solicitud.operadores if op.id_tipo_posicion == tipo_pos]
            posiciones_tipo = [pos for pos in solicitud.posiciones if pos.tipo_posicion == tipo_pos]
            
            if operadores_tipo and posiciones_tipo:
                max_requeridos = max(pos.op_requeridos for pos in posiciones_tipo)
                if len(operadores_tipo) < max_requeridos:
                    advertencias.append(f"Pocos operadores para tipo {tipo_pos}: {len(operadores_tipo)} disponibles, {max_requeridos} requeridos")
        
        return {
            "valido": len(errores) == 0,
            "errores": errores,
            "advertencias": advertencias,
            "estadisticas": {
                "total_operadores": len(solicitud.operadores),
                "total_posiciones": len(solicitud.posiciones),
                "total_ciclos": len(solicitud.ciclos),
                "tipos_posicion": len(tipos_posicion_posiciones)
            }
        }
        
    except Exception as e:
        logger.error(f"Error al validar datos: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al validar datos: {str(e)}")