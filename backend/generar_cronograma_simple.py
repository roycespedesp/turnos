#!/usr/bin/env python3
"""
Generador simple de cronograma de turnos
Sin OR-Tools, solo lógica de ciclos 14x7
"""

import pandas as pd
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Operadores regulares
operadores_regulares = [
    {
        "id_operador": "46781909",
        "nombre": "AGUIRRE HUAYRA JUAN ANTONIO",
        "posicion": "central_1",
        "dia_ciclo_inicial": 19,  # DESCANSO
        "id_cal": 1467,
        "fecha_gen_vac": "25/01/2021",  # Fecha de generación de vacaciones
        "vac_pendientes": 0
    },
    {
        "id_operador": "71023970",
        "nombre": "PEREZ CARDENAS CHRISTIAN DANNY",
        "posicion": "central_1",
        "dia_ciclo_inicial": 12,  # NOCHE
        "id_cal": 1467,
        "fecha_gen_vac": "21/05/2019",  # MÁS ANTIGUA + tiene pendientes
        "vac_pendientes": 30
    },
    {
        "id_operador": "70237797",
        "nombre": "PATRICIO CHAVEZ WALDIR",
        "posicion": "central_1",
        "dia_ciclo_inicial": 5,  # DIA
        "id_cal": 1467,
        "fecha_gen_vac": "21/05/2019",  # MÁS ANTIGUA (misma fecha que PEREZ)
        "vac_pendientes": 0
    },
    {
        "id_operador": "71660619",
        "nombre": "CALIXTO RAMOS ADRIAN",
        "posicion": "central_2",
        "dia_ciclo_inicial": 19,  # DESCANSO
        "id_cal": 1467,
        "fecha_gen_vac": "18/12/2024",  # MÁS RECIENTE
        "vac_pendientes": 0
    },
    {
        "id_operador": "46539662",
        "nombre": "TENORIO TENORIO ROSSMELL JAVIER",
        "posicion": "central_2",
        "dia_ciclo_inicial": 5,  # DIA (ajustado)
        "id_cal": 1467,
        "fecha_gen_vac": "22/02/2024",
        "vac_pendientes": 0
    },
    {
        "id_operador": "70431736",
        "nombre": "HUARCAYA CORDOVA SECILIO MARCELINO",
        "posicion": "central_2",
        "dia_ciclo_inicial": 12,  # NOCHE
        "id_cal": 1467,
        "fecha_gen_vac": "03/07/2024",
        "vac_pendientes": 0
    }
]

# Operador de reemplazo (vacaciones)
operador_reemplazo = {
    "id_operador": "70320115",
    "nombre": "SARMIENTO ZACARIAS CRISTIAN FRANK",
    "posicion": "op_vacaciones",  # Posición abstracta
    "dia_ciclo_inicial": 1,  # DIA
    "id_cal": 1467,
    "fecha_gen_vac": "13/03/2024",
    "vac_pendientes": 0
}

# Todos los operadores
operadores = operadores_regulares + [operador_reemplazo]

# Configuración del ciclo 14x7
CICLO_ID = "14x7"
DIAS_DIA = 7  # Días 1-7: turno DÍA
DIAS_NOCHE = 7  # Días 8-14: turno NOCHE
DIAS_DESCANSO = 7  # Días 15-21: DESCANSO
TOTAL_CICLO = 21
HORAS_POR_TURNO = 12

def calcular_estado_ciclo(dia_ciclo):
    """
    Determina el estado según el día del ciclo 14x7
    Días 1-7: DÍA
    Días 8-14: NOCHE
    Días 15-21: DESCANSO
    """
    if 1 <= dia_ciclo <= 7:
        return "t.dia", "TD"
    elif 8 <= dia_ciclo <= 14:
        return "t.noche", "TN"
    else:  # 15-21
        return "descansando", "DE"

def convertir_fecha(fecha_str):
    """Convierte fecha DD/MM/YYYY a objeto date"""
    try:
        partes = fecha_str.split('/')
        return date(int(partes[2]), int(partes[1]), int(partes[0]))
    except:
        return date(2099, 12, 31)  # Fecha lejana por defecto

def generar_cronograma():
    """Genera el cronograma completo para el año 2025 con bloques vacacionales dinámicos"""
    
    inicio = date(2025, 1, 1)
    fin = date(2025, 12, 31)
    dias_totales = (fin - inicio).days + 1
    
    cronograma = []
    
    # Ordenar operadores regulares por prioridad de vacaciones
    # 1. vac_pendientes > 0 (URGENTE)
    # 2. fecha_gen_vac más antigua (vencimiento cercano)
    operadores_ordenados = sorted(
        operadores_regulares,
        key=lambda op: (
            0 if op.get("vac_pendientes", 0) > 0 else 1,  # Pendientes primero
            convertir_fecha(op.get("fecha_gen_vac", "31/12/2099"))  # Más antiguo primero
        )
    )
    
    print("\n>> ORDEN DE VACACIONES (por urgencia):")
    for idx, op in enumerate(operadores_ordenados, 1):
        fecha_venc = convertir_fecha(op.get("fecha_gen_vac", ""))
        pend_str = f"[PENDIENTES: {op['vac_pendientes']} dias]" if op.get("vac_pendientes", 0) > 0 else ""
        print(f"  {idx}. {op['nombre'][:40]:40s} - Venc: {fecha_venc.strftime('%d/%m/%Y')} {pend_str}")
    print()
    
    # Estado de cada operador
    estado_operadores = {}
    for op in operadores:
        estado_operadores[op["id_operador"]] = {
            "horas_ano": 0,
            "horas_dia_acumuladas": 0,
            "horas_noche_acumuladas": 0,
            "dia_ciclo_actual": op["dia_ciclo_inicial"],
            "en_vacaciones": False,
            "dia_vacacion": 0,
            "necesita_vacaciones": op["id_operador"] in [o["id_operador"] for o in operadores_regulares],
            "prioridad_vacaciones": next((i for i, o in enumerate(operadores_ordenados) if o["id_operador"] == op["id_operador"]), 999)
        }
        
    # Para cada día del año
    for dia_numero in range(dias_totales):
        fecha_actual = inicio + timedelta(days=dia_numero)
        
        # PASO 0: Activar vacaciones DINÁMICAMENTE (solo al terminar ciclo completo)
        # Buscar el siguiente operador en la cola de vacaciones que esté listo
        operadores_pendientes_vacaciones = [
            (op_id, est) for op_id, est in estado_operadores.items()
            if est.get("necesita_vacaciones") and not est.get("en_vacaciones")
        ]
        
        if operadores_pendientes_vacaciones:
            # Ordenar por prioridad (el de menor prioridad va primero)
            operadores_pendientes_vacaciones.sort(key=lambda x: x[1].get("prioridad_vacaciones", 999))
            
            # Verificar si el siguiente en la cola está en período de descanso
            siguiente_id, siguiente_estado = operadores_pendientes_vacaciones[0]
            dia_ciclo = siguiente_estado["dia_ciclo_actual"]
            
            # ✅ CRÍTICO: Solo activar vacaciones si está EN DESCANSO (días 15-21 del ciclo)
            if 15 <= dia_ciclo <= 21:
                # Verificar que nadie más esté de vacaciones actualmente
                hay_alguien_de_vacaciones = any(
                    est.get("en_vacaciones") for op_id, est in estado_operadores.items()
                    if op_id != operador_reemplazo["id_operador"]
                )
                
                if not hay_alguien_de_vacaciones:
                    # ✅ ACTIVAR VACACIONES
                    siguiente_estado["en_vacaciones"] = True
                    siguiente_estado["dia_vacacion"] = 1
                    siguiente_estado["necesita_vacaciones"] = False  # Ya no está pendiente
                    
                    # Obtener info del operador
                    operador_info = next((op for op in operadores_regulares if op["id_operador"] == siguiente_id), None)
                    if operador_info:
                        fecha_fin_vacaciones = fecha_actual + timedelta(days=29)
                        print(f">> {operador_info['nombre']} inicia vacaciones el {fecha_actual.strftime('%d/%m/%Y')}")
                        print(f"   Periodo: {fecha_actual.strftime('%d/%m/%Y')} -> {fecha_fin_vacaciones.strftime('%d/%m/%Y')} (30 dias)")
                        print(f"   Posicion a cubrir: {operador_info['posicion']}")
                        print(f"   Inicia al terminar su ciclo (dia {dia_ciclo}/21)\n")
        
        # Buscar qué operador está de vacaciones hoy
        operador_vacacionista = None
        for op in operadores_regulares:
            estado = estado_operadores.get(op["id_operador"])
            if estado and estado.get("en_vacaciones"):
                operador_vacacionista = op
                break
        
        # Para cada operador
        for operador in operadores:
            es_reemplazo = operador["id_operador"] == operador_reemplazo["id_operador"]
            estado = estado_operadores[operador["id_operador"]]
            dia_ciclo_actual = estado["dia_ciclo_actual"]
            
            # Si este operador está de vacaciones
            if operador_vacacionista and operador["id_operador"] == operador_vacacionista["id_operador"]:
                dia_vacacion = estado["dia_vacacion"]
                
                registro = {
                    "Fecha": fecha_actual.strftime("%d/%m/%Y"),
                    "ID_Operador": operador["id_operador"],
                    "Nombre": operador["nombre"],
                    "Posición": "vacaciones",
                    "Estado": "vacaciones",
                    "Estado2": "VC",
                    "Ciclo": CICLO_ID,
                    "Día_Ciclo": f"'{dia_vacacion}/30",
                    "Desc_Pend": 0,
                    "Vac_Pend": max(0, 30 - dia_vacacion),
                    "Horas_Ciclo": 0,
                    "Horas_Año": estado["horas_ano"],
                    "Horas_Día": estado["horas_dia_acumuladas"],
                    "%_Día": round((estado["horas_dia_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Horas_Noche": estado["horas_noche_acumuladas"],
                    "%_Noche": round((estado["horas_noche_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Observaciones": f"Vacaciones - Dia {dia_vacacion}/30"
                }
                cronograma.append(registro)
                continue
            
            # Si es el operador de reemplazo y hay alguien de vacaciones
            if es_reemplazo and operador_vacacionista:
                # Calcular estado según día del ciclo (ciclo 12x9)
                estado_turno, estado2 = calcular_estado_ciclo(dia_ciclo_actual)
                
                # Calcular horas del día
                if estado_turno == "t.dia":
                    horas_dia = HORAS_POR_TURNO
                    horas_noche = 0
                    estado["horas_dia_acumuladas"] += HORAS_POR_TURNO
                    estado["horas_ano"] += HORAS_POR_TURNO
                elif estado_turno == "t.noche":
                    horas_dia = 0
                    horas_noche = HORAS_POR_TURNO
                    estado["horas_noche_acumuladas"] += HORAS_POR_TURNO
                    estado["horas_ano"] += HORAS_POR_TURNO
                else:  # descansando
                    horas_dia = 0
                    horas_noche = 0
                
                # Calcular horas del ciclo
                if dia_ciclo_actual == 1:
                    horas_ciclo = horas_dia + horas_noche
                else:
                    dias_transcurridos_ciclo = dia_ciclo_actual - 1
                    if dias_transcurridos_ciclo <= 7:
                        horas_ciclo = dias_transcurridos_ciclo * 12 + horas_dia
                    elif dias_transcurridos_ciclo <= 14:
                        horas_ciclo = 7 * 12 + (dias_transcurridos_ciclo - 7) * 12 + horas_noche
                    else:
                        horas_ciclo = 14 * 12
                
                # POSICIÓN REAL que está cubriendo (no op_vacaciones)
                posicion = operador_vacacionista["posicion"] if estado_turno != "descansando" else "descanso"
                
                registro = {
                    "Fecha": fecha_actual.strftime("%d/%m/%Y"),
                    "ID_Operador": operador["id_operador"],
                    "Nombre": operador["nombre"],
                    "Posición": posicion,
                    "Estado": estado_turno,
                    "Estado2": estado2,
                    "Ciclo": CICLO_ID,
                    "Día_Ciclo": f"'{dia_ciclo_actual}/21",
                    "Desc_Pend": 0,
                    "Vac_Pend": 0,
                    "Horas_Ciclo": horas_ciclo,
                    "Horas_Año": estado["horas_ano"],
                    "Horas_Día": estado["horas_dia_acumuladas"],
                    "%_Día": round((estado["horas_dia_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Horas_Noche": estado["horas_noche_acumuladas"],
                    "%_Noche": round((estado["horas_noche_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Observaciones": f"Reemplaza a {operador_vacacionista['nombre']}"
                }
                cronograma.append(registro)
                
                # Avanzar ciclo
                estado["dia_ciclo_actual"] += 1
                if estado["dia_ciclo_actual"] > TOTAL_CICLO:
                    estado["dia_ciclo_actual"] = 1
                continue
            
            # Si es el operador de reemplazo pero NO hay nadie de vacaciones
            if es_reemplazo and not operador_vacacionista:
                # No trabaja
                registro = {
                    "Fecha": fecha_actual.strftime("%d/%m/%Y"),
                    "ID_Operador": operador["id_operador"],
                    "Nombre": operador["nombre"],
                    "Posición": "descanso",
                    "Estado": "descansando",
                    "Estado2": "DE",
                    "Ciclo": CICLO_ID,
                    "Día_Ciclo": f"'0/21",
                    "Desc_Pend": 0,
                    "Vac_Pend": 0,
                    "Horas_Ciclo": 0,
                    "Horas_Año": estado["horas_ano"],
                    "Horas_Día": estado["horas_dia_acumuladas"],
                    "%_Día": round((estado["horas_dia_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Horas_Noche": estado["horas_noche_acumuladas"],
                    "%_Noche": round((estado["horas_noche_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                    "Observaciones": "Sin reemplazos asignados"
                }
                cronograma.append(registro)
                continue
            
            # Operador regular (no de vacaciones, no vacacionista)
            # Calcular estado según día del ciclo
            estado_turno, estado2 = calcular_estado_ciclo(dia_ciclo_actual)
            
            # Calcular horas del día
            if estado_turno == "t.dia":
                horas_dia = HORAS_POR_TURNO
                horas_noche = 0
                estado["horas_dia_acumuladas"] += HORAS_POR_TURNO
                estado["horas_ano"] += HORAS_POR_TURNO
            elif estado_turno == "t.noche":
                horas_dia = 0
                horas_noche = HORAS_POR_TURNO
                estado["horas_noche_acumuladas"] += HORAS_POR_TURNO
                estado["horas_ano"] += HORAS_POR_TURNO
            else:  # descansando
                horas_dia = 0
                horas_noche = 0
            
            # Calcular horas del ciclo
            if dia_ciclo_actual == 1:
                horas_ciclo = horas_dia + horas_noche
            else:
                dias_transcurridos_ciclo = dia_ciclo_actual - 1
                if dias_transcurridos_ciclo <= 7:
                    horas_ciclo = dias_transcurridos_ciclo * 12 + horas_dia
                elif dias_transcurridos_ciclo <= 14:
                    horas_ciclo = 7 * 12 + (dias_transcurridos_ciclo - 7) * 12 + horas_noche
                else:
                    horas_ciclo = 14 * 12
            
            # Determinar posición
            posicion = operador["posicion"] if estado_turno != "descansando" else "descanso"
            
            registro = {
                "Fecha": fecha_actual.strftime("%d/%m/%Y"),
                "ID_Operador": operador["id_operador"],
                "Nombre": operador["nombre"],
                "Posición": posicion,
                "Estado": estado_turno,
                "Estado2": estado2,
                "Ciclo": CICLO_ID,
                "Día_Ciclo": f"'{dia_ciclo_actual}/21",
                "Desc_Pend": 0,
                "Vac_Pend": 0,
                "Horas_Ciclo": horas_ciclo,
                "Horas_Año": estado["horas_ano"],
                "Horas_Día": estado["horas_dia_acumuladas"],
                "%_Día": round((estado["horas_dia_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                "Horas_Noche": estado["horas_noche_acumuladas"],
                "%_Noche": round((estado["horas_noche_acumuladas"] / estado["horas_ano"] * 100), 2) if estado["horas_ano"] > 0 else 0,
                "Observaciones": ""
            }
            cronograma.append(registro)
            
            # Avanzar ciclo o día de vacación
            if estado_operadores[operador["id_operador"]].get("en_vacaciones"):
                estado_operadores[operador["id_operador"]]["dia_vacacion"] += 1
                
                # Validar si completó los 30 días de vacaciones
                if estado_operadores[operador["id_operador"]]["dia_vacacion"] >= 30:
                    dias_tomados = estado_operadores[operador["id_operador"]]["dia_vacacion"]
                    estado_operadores[operador["id_operador"]]["en_vacaciones"] = False
                    estado_operadores[operador["id_operador"]]["dia_vacacion"] = 0
                    estado_operadores[operador["id_operador"]]["dia_ciclo_actual"] = 1  # Reiniciar ciclo
                    print(f">> {operador['nombre']} termino vacaciones el {fecha_actual.strftime('%d/%m/%Y')} - Total: {dias_tomados} dias\n")
            else:
                estado["dia_ciclo_actual"] += 1
                if estado["dia_ciclo_actual"] > TOTAL_CICLO:
                    estado["dia_ciclo_actual"] = 1
    
    return cronograma

def generar_vista_calendario(df):
    """Genera vista calendario: operadores en filas, días en columnas"""
    
    # Crear estructura para calendario con separadores entre meses
    inicio = date(2025, 1, 1)
    fin = date(2025, 12, 31)
    
    # Preparar datos
    calendario_data = []
    
    for operador in operadores:
        fila = {
            "Operador": operador["nombre"],
            "Posición": operador["posicion"]
        }
        
        # Filtrar datos del operador
        df_operador = df[df["Nombre"] == operador["nombre"]]
        
        # Para cada día del año
        fecha_actual = inicio
        mes_anterior = 0
        
        while fecha_actual <= fin:
            # Si cambió de mes, agregar columna separadora
            if fecha_actual.month != mes_anterior and mes_anterior != 0:
                col_sep = f"SEP_{mes_anterior}"
                fila[col_sep] = ""
            
            mes_anterior = fecha_actual.month
            
            # Buscar estado del operador en esa fecha
            fecha_str = fecha_actual.strftime("%d/%m/%Y")
            registro = df_operador[df_operador["Fecha"] == fecha_str]
            
            if not registro.empty:
                estado2 = registro.iloc[0]["Estado2"]
            else:
                estado2 = ""
            
            # Agregar columna con formato: MES_DIA (para ordenar correctamente)
            col_nombre = f"{fecha_actual.month:02d}_{fecha_actual.day:02d}"
            fila[col_nombre] = estado2
            
            fecha_actual += timedelta(days=1)
        
        calendario_data.append(fila)
    
    return pd.DataFrame(calendario_data)

def generar_vista_programacion(df):
    """Genera vista de programación: meses en filas, días en columnas, ID de operador en celdas
    CON ENCABEZADOS REPETIDOS Y SEPARACIÓN POR POSICIÓN"""
    
    # Mapeo de operadores a números
    operadores_map = {
        "AGUIRRE HUAYRA JUAN ANTONIO": 1,
        "PEREZ CARDENAS CHRISTIAN DANNY": 2,
        "PATRICIO CHAVEZ WALDIR": 3,
        "CALIXTO RAMOS ADRIAN": 4,
        "TENORIO TENORIO ROSSMELL JAVIER": 5,
        "HUARCAYA CORDOVA SECILIO MARCELINO": 6,
        "SARMIENTO ZACARIAS CRISTIAN FRANK": 7
    }
    
    # Posiciones disponibles
    posiciones_lista = ["central_1", "central_2", "bocatoma_1", "bocatoma_2"]
    
    # Meses
    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    dias_por_mes = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    
    # Crear estructura de columnas de días
    columnas_dias = []
    for dia in range(1, 32):  # Máximo 31 días
        columnas_dias.append(f"D{dia}")
    
    programacion_data = []
    
    # Generar para CADA posición
    for idx_pos, posicion in enumerate(posiciones_lista):
        # SOLO agregar encabezado si NO es la primera posición
        # (la primera usa el header de pandas)
        if idx_pos > 0:
            # Primero agregar fila vacía de separación
            fila_vacia = {"Posición": "", "Mes": "", "Turno": ""}
            for dia in range(1, 32):
                fila_vacia[f"D{dia}"] = ""
            programacion_data.append(fila_vacia)
            
            # Luego agregar encabezado repetido
            header_row = {"Posición": "Posición", "Mes": "Mes", "Turno": "Turno"}
            for dia in range(1, 32):
                header_row[f"D{dia}"] = dia if dia <= 31 else ""
            programacion_data.append(header_row)
        
        # Crear filas para cada mes (TD y TN)
        for mes_idx, (mes_nombre, dias_mes) in enumerate(zip(meses_nombres, dias_por_mes), start=1):
            for turno in ["TD", "TN"]:
                fila = {
                    "Posición": posicion,
                    "Mes": mes_nombre,
                    "Turno": turno
                }
                
                # Para cada día del mes
                for dia in range(1, 32):
                    if dia <= dias_mes:
                        # Construir fecha
                        fecha_actual = date(2025, mes_idx, dia)
                        fecha_str = fecha_actual.strftime("%d/%m/%Y")
                        
                        # Buscar operador en esa posición/fecha/turno
                        estado_buscar = "t.dia" if turno == "TD" else "t.noche"
                        
                        registro = df[
                            (df["Posición"] == posicion) &
                            (df["Fecha"] == fecha_str) &
                            (df["Estado"] == estado_buscar)
                        ]
                        
                        if not registro.empty:
                            nombre_operador = registro.iloc[0]["Nombre"]
                            numero_operador = operadores_map.get(nombre_operador, 0)
                        else:
                            numero_operador = 0
                        
                        fila[f"D{dia}"] = numero_operador
                    else:
                        # Día no existe en este mes
                        fila[f"D{dia}"] = ""
                
                programacion_data.append(fila)
    
    df_prog = pd.DataFrame(programacion_data)
    
    # Agregar hoja de leyenda como DataFrame separado
    leyenda_data = []
    for nombre, numero in sorted(operadores_map.items(), key=lambda x: x[1]):
        leyenda_data.append({"Número": numero, "Operador": nombre})
    
    df_leyenda = pd.DataFrame(leyenda_data)
    
    return df_prog, df_leyenda

def aplicar_formato_excel(archivo):
    """Aplica formato avanzado a las pestañas del Excel"""
    
    wb = load_workbook(archivo)
    
    # ===== PESTAÑA CRONOGRAMA =====
    if "Cronograma" in wb.sheetnames:
        ws_crono = wb["Cronograma"]
        
        # Congelar paneles (fila 1)
        ws_crono.freeze_panes = "A2"
        
        # Aplicar autofiltro
        ws_crono.auto_filter.ref = ws_crono.dimensions
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_crono[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print("[OK] Formato aplicado a pestaña 'Cronograma'")
    
    # ===== PESTAÑA CALENDARIO =====
    if "Calendario" in wb.sheetnames:
        ws_cal = wb["Calendario"]
        
        # INSERTAR FILA ADICIONAL PARA ENCABEZADO DE MESES
        ws_cal.insert_rows(1)
        
        # Colores más oscuros y suaves
        color_td = PatternFill(start_color="5D8F3E", end_color="5D8F3E", fill_type="solid")  # Verde oscuro
        color_tn = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")  # Azul oscuro
        color_de = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")  # Rojo oscuro
        color_vc = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")  # Naranja oscuro
        
        # Meses en español
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        dias_por_mes = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 2025 no es bisiesto
        
        # Formatear encabezados de columnas fijas
        ws_cal['A1'] = ""
        ws_cal['B1'] = ""
        ws_cal['A2'] = "Operador"
        ws_cal['B2'] = "Posición"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell_ref in ['A2', 'B2']:
            ws_cal[cell_ref].fill = header_fill
            ws_cal[cell_ref].font = header_font
            ws_cal[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        
        # Construir encabezados de meses y días
        col_idx = 3  # Empezar después de Operador y Posición
        
        for mes_num, (mes_nombre, dias_mes) in enumerate(zip(meses_nombres, dias_por_mes), start=1):
            # Fusionar celdas para el nombre del mes
            start_col = col_idx
            end_col = col_idx + dias_mes - 1
            
            # Escribir nombre del mes en la primera fila
            ws_cal.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            cell = ws_cal.cell(row=1, column=start_col)
            cell.value = mes_nombre
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir números de días en la segunda fila
            for dia in range(1, dias_mes + 1):
                cell = ws_cal.cell(row=2, column=col_idx)
                cell.value = dia
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1
            
            # Agregar columna separadora (excepto después del último mes)
            if mes_num < 12:
                cell = ws_cal.cell(row=1, column=col_idx)
                cell.value = ""
                cell = ws_cal.cell(row=2, column=col_idx)
                cell.value = ""
                col_idx += 1
        
        # Aplicar colores a las celdas de datos según el código
        for row_idx, row in enumerate(ws_cal.iter_rows(min_row=3), start=3):
            for cell in row[2:]:  # Saltar columnas Operador y Posición
                if cell.value == "TD":
                    cell.fill = color_td
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "TN":
                    cell.fill = color_tn
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "DE":
                    cell.fill = color_de
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                elif cell.value == "VC":
                    cell.fill = color_vc
                    cell.font = Font(bold=True, color="FFFFFF", size=8)
                
                # Centrar contenido
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Congelar paneles (columnas A-B y filas 1-2)
        ws_cal.freeze_panes = "C3"
        
        # Ajustar ancho de columnas
        ws_cal.column_dimensions['A'].width = 30  # Operador
        ws_cal.column_dimensions['B'].width = 12  # Posición
        
        # Columnas de días: ancho pequeño (3 caracteres)
        for col_idx in range(3, ws_cal.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_cal.column_dimensions[col_letter].width = 3
        
        # Ajustar altura de filas
        ws_cal.row_dimensions[1].height = 20
        ws_cal.row_dimensions[2].height = 18
        
        print("[OK] Formato aplicado a pestaña 'Calendario' con colores oscuros")
    
    # ===== PESTAÑA PROGRAMACIÓN =====
    if "Programacion" in wb.sheetnames:
        ws_prog = wb["Programacion"]
        
        # Paleta de colores por operador (más vibrantes pero no chillones)
        colores_operadores = {
            1: PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid"),  # Amarillo suave
            2: PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),  # Naranja suave
            3: PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid"),  # Verde suave
            4: PatternFill(start_color="6FA8DC", end_color="6FA8DC", fill_type="solid"),  # Azul suave
            5: PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid"),  # Morado suave
            6: PatternFill(start_color="E06666", end_color="E06666", fill_type="solid"),  # Rojo suave
            7: PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid"),  # Naranja claro (Sarmiento)
            0: PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")   # Gris (vacío)
        }
        
        # Formatear encabezados (filas donde columna A = "Posición")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Recorrer todas las filas para detectar encabezados y aplicar formatos
        for row_idx, row in enumerate(ws_prog.iter_rows(min_row=1), start=1):
            # Detectar si es fila de encabezado
            if row[0].value == "Posición":
                # Formatear toda la fila como encabezado
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Formatear filas de datos
                # Aplicar colores según número de operador (columnas D en adelante)
                for cell in row[3:]:  # Saltar Posición, Mes, Turno
                    if isinstance(cell.value, int):
                        cell.fill = colores_operadores.get(cell.value, colores_operadores[0])
                        cell.font = Font(bold=True, color="000000", size=9)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.value == "":
                        # Celdas vacías (días que no existen en el mes)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Formatear columnas fijas (Posición, Mes, Turno)
                for col in range(0, 3):
                    cell = row[col]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Congelar paneles (solo columnas A-C, NO filas)
        ws_prog.freeze_panes = "D1"
        
        # Ajustar anchos
        ws_prog.column_dimensions['A'].width = 15  # Posición
        ws_prog.column_dimensions['B'].width = 12  # Mes
        ws_prog.column_dimensions['C'].width = 8   # Turno
        
        # Días: ancho pequeño
        for col_idx in range(4, ws_prog.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_prog.column_dimensions[col_letter].width = 3
        
        # NO APLICAR AUTOFILTRO
        
        print("[OK] Formato aplicado a pestaña 'Programacion' con colores por operador (sin filtros)")
    
    # ===== PESTAÑA LEYENDA =====
    if "Leyenda" in wb.sheetnames:
        ws_ley = wb["Leyenda"]
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_ley[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar colores a la columna Número según operador
        for row_idx, row in enumerate(ws_ley.iter_rows(min_row=2), start=2):
            numero_cell = ws_ley.cell(row=row_idx, column=1)
            if isinstance(numero_cell.value, int):
                numero_cell.fill = colores_operadores.get(numero_cell.value, colores_operadores[0])
                numero_cell.font = Font(bold=True, color="000000", size=11)
                numero_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formatear nombre del operador
            nombre_cell = ws_ley.cell(row=row_idx, column=2)
            nombre_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar anchos
        ws_ley.column_dimensions['A'].width = 10
        ws_ley.column_dimensions['B'].width = 35
        
        print("[OK] Formato aplicado a pestaña 'Leyenda'")
    
    # Guardar cambios
    wb.save(archivo)
    print(f"[OK] Formato guardado en: {archivo}")

def main():
    """Función principal"""
    print("=" * 70)
    print("GENERANDO CRONOGRAMA CON 4 PESTANAS Y BLOQUES VACACIONALES DINAMICOS")
    print("=" * 70)
    print(f"Periodo: 01/01/2025 - 31/12/2025")
    print(f"Operadores regulares: {len(operadores_regulares)}")
    print(f"Operador de reemplazo: 1 (Sarmiento)")
    print(f"Total operadores: {len(operadores)}")
    print()
    print("PRIORIDAD DE VACACIONES:")
    print("  - Ordenadas por fecha_gen_vac (más antiguas = mayor urgencia)")
    print("  - Los que tienen vac_pendientes > 0 van primero")
    print("  - Se activan dinámicamente cuando cada operador termine su ciclo (día 15-21)")
    print("  - Total: 180 días continuos de vacaciones (30 días x 6 operadores)")
    print()
    
    # Generar cronograma
    cronograma = generar_cronograma()
    
    print(f"[OK] Cronograma generado: {len(cronograma)} registros")
    print()
    
    # Crear DataFrame
    df = pd.DataFrame(cronograma)
    
    # Mostrar estadísticas
    turnos_dia = len(df[df["Estado"] == "t.dia"])
    turnos_noche = len(df[df["Estado"] == "t.noche"])
    dias_descanso = len(df[df["Estado"] == "descansando"])
    
    print("ESTADISTICAS:")
    print(f"   - Turnos DIA: {turnos_dia}")
    print(f"   - Turnos NOCHE: {turnos_noche}")
    print(f"   - Dias DESCANSO: {dias_descanso}")
    print()
    
    # Generar vista calendario
    print("[...] Generando vista calendario...")
    df_calendario = generar_vista_calendario(df)
    print(f"[OK] Vista calendario generada: {len(df_calendario)} operadores x 365 dias")
    print()
    
    # Generar vista programación
    print("[...] Generando vista programacion...")
    df_programacion, df_leyenda = generar_vista_programacion(df)
    print(f"[OK] Vista programacion generada: {len(df_programacion)} filas (meses x turnos x posiciones)")
    print(f"[OK] Leyenda generada: {len(df_leyenda)} operadores")
    print()
    
    # Exportar a Excel con 4 pestañas
    archivo = "cronograma_2025_simple.xlsx"
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Cronograma", index=False)
        df_calendario.to_excel(writer, sheet_name="Calendario", index=False)
        df_programacion.to_excel(writer, sheet_name="Programacion", index=False)
        df_leyenda.to_excel(writer, sheet_name="Leyenda", index=False)
    
    print(f"[OK] Archivo base exportado: {archivo}")
    print()
    
    # Aplicar formato avanzado
    print("[...] Aplicando formato (colores, filtros, paneles fijos)...")
    aplicar_formato_excel(archivo)
    print()
    
    # Mostrar muestra
    print("MUESTRA CRONOGRAMA (primeros 10 registros):")
    print(df.head(10)[["Fecha", "Nombre", "Posición", "Estado", "Estado2", "Día_Ciclo"]].to_string())
    print()
    print("=" * 70)
    print(f"ARCHIVO GENERADO: {archivo}")
    print("PESTANAS:")
    print("  1. Cronograma - Datos detallados (SIN Merge1, Merge2, Prog)")
    print("  2. Calendario - Vista visual por operador (TD/TN/DE)")
    print("  3. Programacion - Vista mensual por posicion (numero de operador)")
    print("  4. Leyenda - Mapeo numero -> operador")
    print("=" * 70)
    
    return archivo

if __name__ == "__main__":
    main()

